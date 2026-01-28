
import openpyxl
import os

DATA_FILE = "/app/data/Active and Inconclusive Properties.xlsx"

def inspect():
    if not os.path.exists(DATA_FILE):
        print("File not found.")
        return

    print(f"Loading workbook (read_only=True)...")
    wb = openpyxl.load_workbook(DATA_FILE, read_only=True)
    sheet = wb.active
    
    print("Reading headers...")
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    print(f"Headers: {headers}")
    
    # Estimate row count or just confirm matches
    count = 0
    ct_count = 0
    state_col_idx = None
    
    # Find State column index
    try:
        state_col_idx = headers.index('State')
    except ValueError:
        print("State column not found in headers!")
    
    if state_col_idx is not None:
        print("Scanning rows for CT properties...")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            count += 1
            if count % 5000 == 0:
                print(f"Scanned {count} rows...")
            
            if row[state_col_idx] == 'CT':
                ct_count += 1
                if ct_count < 3:
                     print(f"Sample CT Row: {row}")

    print(f"Total Rows: {count}")
    print(f"CT Rows: {ct_count}")

if __name__ == "__main__":
    inspect()
